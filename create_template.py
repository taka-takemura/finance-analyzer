# -*- coding: utf-8 -*-
"""入力テンプレート・サンプルデータの Excel を生成する。

usage: python create_template.py
→ テンプレート.xlsx / サンプルデータ.xlsx を生成
"""
import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

YEARS = ["2021年度", "2022年度", "2023年度", "2024年度", "2025年度"]

PL_ITEMS = ["売上高", "売上原価", "売上総利益", "販売費及び一般管理費",
            "人件費", "減価償却費", "営業利益", "営業外収益", "営業外費用",
            "支払利息", "経常利益", "特別利益", "特別損失",
            "税引前当期純利益", "法人税等", "当期純利益"]
BS_ITEMS = ["現金及び預金", "売上債権", "棚卸資産", "その他流動資産", "流動資産合計",
            "有形固定資産", "無形固定資産", "投資その他の資産", "固定資産合計",
            "資産合計", "仕入債務", "短期借入金", "その他流動負債", "流動負債合計",
            "長期借入金", "その他固定負債", "固定負債合計", "負債合計", "純資産合計",
            "自己資本", "非支配株主持分"]
CF_ITEMS = ["営業活動によるキャッシュ・フロー", "投資活動によるキャッシュ・フロー",
            "財務活動によるキャッシュ・フロー", "現金及び現金同等物の期末残高", "設備投資額"]


def _style(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
    ws.column_dimensions["A"].width = 34
    for i in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    for row in ws.iter_rows(min_row=2, min_col=2):
        for c in row:
            c.number_format = "#,##0"
    ws.freeze_panes = "B2"


def write_book(path: str, data: dict):
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        for sheet, df in data.items():
            df.to_excel(w, sheet_name=sheet, index_label="科目")
            _style(w.sheets[sheet])
    print(f"生成: {path}")


def make_template():
    data = {
        "PL": pd.DataFrame(index=PL_ITEMS, columns=YEARS),
        "BS": pd.DataFrame(index=BS_ITEMS, columns=YEARS),
        "CF": pd.DataFrame(index=CF_ITEMS, columns=YEARS),
    }
    write_book("テンプレート.xlsx", data)


def make_sample():
    """架空の製造業(単位:百万円)。5年で成長・原価率改善のストーリー。"""
    rng = np.random.default_rng(42)
    n = len(YEARS)
    sales = np.array([48_000, 51_500, 54_200, 58_800, 63_500], dtype=float)
    cogs_ratio = np.array([0.66, 0.65, 0.645, 0.63, 0.62])
    cogs = sales * cogs_ratio
    gross = sales - cogs
    sga = sales * np.array([0.235, 0.232, 0.23, 0.225, 0.222])
    op = gross - sga
    noi = sales * 0.004
    noe = np.array([420, 410, 430, 390, 370], dtype=float)
    interest = noe * 0.7
    ordinary = op + noi - noe
    extra_g = np.array([0, 150, 0, 80, 0], dtype=float)
    extra_l = np.array([120, 0, 260, 0, 90], dtype=float)
    pretax = ordinary + extra_g - extra_l
    tax = pretax * 0.30
    ni = pretax - tax

    pl = pd.DataFrame({
        "売上高": sales, "売上原価": cogs, "売上総利益": gross,
        "販売費及び一般管理費": sga, "人件費": sga * 0.55,
        "減価償却費": sales * 0.035, "営業利益": op,
        "営業外収益": noi, "営業外費用": noe, "支払利息": interest,
        "経常利益": ordinary, "特別利益": extra_g, "特別損失": extra_l,
        "税引前当期純利益": pretax, "法人税等": tax, "当期純利益": ni,
    }, index=YEARS).T

    cash = np.array([6_200, 6_900, 7_400, 8_600, 10_200], dtype=float)
    recv = sales * np.array([0.19, 0.19, 0.185, 0.18, 0.175])
    inv = sales * np.array([0.135, 0.13, 0.128, 0.122, 0.118])
    oca = sales * 0.02
    ca = cash + recv + inv + oca
    ppe = np.array([16_500, 16_900, 17_600, 18_400, 19_300], dtype=float)
    intan = np.array([1_800, 1_750, 1_900, 2_100, 2_250], dtype=float)
    invest = np.array([3_200, 3_300, 3_400, 3_600, 3_800], dtype=float)
    fa = ppe + intan + invest
    assets = ca + fa
    pay = sales * np.array([0.115, 0.113, 0.112, 0.11, 0.108])
    std = np.array([2_800, 2_600, 2_500, 2_200, 2_000], dtype=float)
    ocl = sales * 0.04
    cl = pay + std + ocl
    ltd = np.array([7_500, 7_000, 6_800, 6_200, 5_600], dtype=float)
    ofl = sales * 0.025
    fl = ltd + ofl
    liab = cl + fl
    equity = assets - liab
    minority = (equity * 0.02).round(0)   # 非支配株主持分 2%
    own_equity = equity - minority        # 自己資本(親会社帰属)

    bs = pd.DataFrame({
        "現金及び預金": cash, "売上債権": recv, "棚卸資産": inv,
        "その他流動資産": oca, "流動資産合計": ca,
        "有形固定資産": ppe, "無形固定資産": intan, "投資その他の資産": invest,
        "固定資産合計": fa, "資産合計": assets,
        "仕入債務": pay, "短期借入金": std, "その他流動負債": ocl,
        "流動負債合計": cl, "長期借入金": ltd, "その他固定負債": ofl,
        "固定負債合計": fl, "負債合計": liab, "純資産合計": equity,
        "自己資本": own_equity, "非支配株主持分": minority,
    }, index=YEARS).T

    ocf = ni + sales * 0.035 + rng.normal(300, 100, n)
    capex = -(ppe * 0.12 + rng.normal(200, 50, n))
    icf = capex - 150
    fcf_fin = np.array([-900, -1_100, -800, -1_300, -1_400], dtype=float)
    cf = pd.DataFrame({
        "営業活動によるキャッシュ・フロー": ocf,
        "投資活動によるキャッシュ・フロー": icf,
        "財務活動によるキャッシュ・フロー": fcf_fin,
        "現金及び現金同等物の期末残高": cash,
        "設備投資額": -capex,
    }, index=YEARS).T

    write_book("サンプルデータ.xlsx", {"PL": pl.round(0), "BS": bs.round(0), "CF": cf.round(0)})


if __name__ == "__main__":
    make_template()
    make_sample()
